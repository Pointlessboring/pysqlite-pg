##################################
# Daily SQLite Grind
#

import sqlite3
from sqlite3 import Error

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from datetime import date, datetime, timedelta
from random import randint

#####################################
# Basic date manipulation exercises
#

today = date.today()
print(f"\nToday is: {today.strftime('%A %B %d, the %jth day of %Y')}")

####################################
# Getting names from an excel file
#

print("\nExtracting names from data files.")

try: 
    workbook = load_workbook(filename="FirstLastNames.xlsx")
    sheet = workbook["Surnames"]
 
    lastnames = []
    firstnames = []

    for cell in sheet["A"]:
        lastnames.append(cell.value)
    lastnames.sort()
    
    sheet = workbook["Firstnames"]
    for row in sheet.iter_rows():
       firstnames.append((row[0].value, row[1].value))
    firstnames.sort()
 
except Error as err:
    print(f"Error Reading first/last name dataset: {err}")

print("Collecting Jobs from data files.")

try: 
    workbook = load_workbook(filename="Jobslist.xlsx")
    sheet = workbook["Jobs"]
 
    joblist = []
    for cell in sheet["A"]:
        joblist.append(cell.value)
    joblist.sort()

except Error as err:
    print(f"Error Reading Jobs dataset: {err}")

###############################
# DB processing
#

# Create DB filename
filename = "UselessDB" + today.strftime("%y%m%d")+".db"

# open DB connection

try: 
    conn = sqlite3.Connection(filename)

except Error as err:
    print(err)

# Create cursor into DB
cursor = conn.cursor()

# Get list of tables in the DB
tables = cursor.execute("SELECT name FROM sqlite_schema WHERE type ='table'").fetchall()

# if the people table does not exist, create it.
if tables == [] or (('people',) not in tables):
    cursor.execute("""
                    CREATE TABLE people (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    first CHAR(30),
                    last CHAR(30),
                    birthday DATE,
                    gender CHAR(1)
                    )
                    """)
    print("Creating PEOPLE table into the database.")

if tables == [] or (('jobs',) not in tables):
    cursor.execute("""
                    CREATE TABLE jobs (
                    id INTEGER PRIMARY KEY,
                    title CHAR(100),
                    salary INTEGER
                    )
                    """)
    print("\nCreating JOBS table into the database.")

# Add 1000 records to the people table.
print("\nAdding 1000 entries into the people and jobs tables database.")

for i in range(1,1000):
    firstnameID = randint(0,len(firstnames)-1)
    lastnameID = randint(0,len(lastnames)-1)
    jobnameID = randint(0,len(joblist)-1)

    # Add to people table
    cursor.execute(f"INSERT INTO people (first, last, birthday, gender) VALUES(?,?,?,?)", 
                    (firstnames[firstnameID][0], 
                     lastnames[lastnameID], 
                     date.today()-timedelta(randint(0,365*80)),
                     firstnames[firstnameID][1]
                    ))

    # Add info to jobs table
    newID = cursor.execute('SELECT seq FROM sqlite_sequence WHERE name="people" ').fetchall()

    cursor.execute(f"INSERT INTO jobs (id, title, salary) VALUES (?,?,?)",
                    (newID[0][0],
                    joblist[jobnameID],
                    randint(0,100000) + 20000
                    ))

    conn.commit()


###############################
# Read info back from database
#
print("Collecting all informations from people and jobs tables in the database.")
 
data = cursor.execute("SELECT p.first, p.last, p.birthday, p.gender, j.title, j.salary FROM people p INNER JOIN jobs j ON p.id = j.id").fetchall()
data.sort()

# close the DB connection
conn.close()

################################
# Write data back to excel sheet
#
print("Writing people/jobs information into an Excel workbook.")

wb = Workbook()
ws = wb.active
ws.title = "My Results"

# Inserting the header row
header = ['First', 'Last', 'Birthday', 'Gender', 'Job Type', 'Salary']
for y in range(0,len(header)):
    ws.cell(row=1, column=y+1, value = header[y])
    mycell = ws.cell(row=1, column=y+1)
    mycell.font = Font(bold=True)
    mycell.alignment = Alignment(horizontal="center", vertical="center")
    mycell.fill = PatternFill("solid", fgColor="DDDDDD")

# Formatting the salary column
col = ws.column_dimensions['F']
col.number_format = '0.00'

# Inserting the data
for row in data:
    ws.append(row)

# Adjusting the column width 
for colname in ['A', 'B', 'C', 'D', 'E']:
    maxsize = 0
    for cell in ws[colname]:
        maxsize = max(maxsize, len(cell.value))
    ws.column_dimensions[colname].width = maxsize

# Formatting the data in the cells
for cell in ws['F']:
    cell.number_format = u'_($* #,##0_)'

wb.save('UselessResultsFile.xlsx')

print("End program.\n")


