######################################
# Daily Grind
#
# Todo: Delete the empty rows at the bottom of the CSV files.

import sqlite3, time, csv
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def importdata(mydbname, tablename, csvfilename, label):
    """ Function to import csv file into a DB """    
    print(f"Validating if {label} table exists in the database.")

    # Get list of tables from DB, create them if they are not there.
    
    conn = sqlite3.Connection(mydbname) # open DB Connection
    cursor = conn.cursor()   
    tables = cursor.execute("SELECT name from sqlite_schema").fetchall()

    if tables == [] or (tablename,) not in tables:      

        # importing csv file
        with open(csvfilename) as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=',')

            # csv_reader is an iterator. Converted to a tuple of data to use it more freely.
            csvDB = tuple(csv_reader) 

        # Get csv file info: # of fields, header names, etc.
        fieldnames = csvDB[0]
        nbfields = len(fieldnames)
        fieldnamestring = "'" + "', '".join(fieldnames) + "'"
  
        sqlstring = f"CREATE TABLE {tablename} ("            # Beginning of SQL CREATE command
        for field in fieldnames:
            sqlstring += " '" + field + "' TEXT,"         # Add the various field to create
    
        sqlstring = sqlstring[:-1] + ")"                # Remove last comma and close the bracket
        cursor.execute(sqlstring)

        print(f"Creating {label} table into the database.")

        ###############################################
        # Populate the database with data from CSV file
        #
        # We create a SQLstring command and a variable holding the data to be insert

        print(f"Populating {label} table into the database.\n")

        for row in csvDB[1:]:
            sqlstring = f"INSERT INTO {tablename} ({fieldnamestring}) VALUES ({('?,'*nbfields)[:-1]})"
            cursor.execute(sqlstring, row)
    conn.commit()
    conn.close()

# start timer for performance measurement
start_timer = time.time()

# Greeting message
print(f"\nWelcome! Today is {date.today().strftime('%A %B %d, the %jth day of %Y')}\n")

# Create filenames based on today's date
basename = 'Useless'+date.today().strftime("%y%m%d")
logname = basename + ".log"               # log file
dbname = basename + ".db"                 # DB file
xlname = basename + ".xlsx"               # Excel file

# Begin logging activities
with open(logname, 'a') as f:
    f.write(f'{datetime.now().strftime("%y-%m-%d %X")} Launching Daily grind program...\n')

importdata(dbname, "population", "Datasets/worldbank-population-2021.csv", "POPULATION")
importdata(dbname, "gdp", "Datasets/worldbank-GDP-2021.csv", "GDP")
importdata(dbname, "debt", "Datasets/worldbank-External Debt-2021.csv", "DEBT")


#########
# Perform data treatment here...

# open DB Connection
conn = sqlite3.Connection(dbname)
cursor = conn.cursor()

sqlstring = ('Select t1."Country Name", t1.debtvalue, t2.pop '
            'FROM (SELECT "Country Name", "2021 [YR2021]" as debtvalue from debt) AS t1 '
            'JOIN ( SELECT "Country Name", "2021 [YR2021]" AS pop from population) AS t2 '
            'ON t1."Country Name" = t2."Country Name" ')

data = cursor.execute(sqlstring).fetchall()

#####################################
# Write data back to excel file/sheet
#

print("Saving Data into an Excel workbook.")
wb = Workbook()
ws = wb.active
ws.title = "World Bank Country information"

x = 0
for line in data:
    x += 1
    y = 0
    for element in line:
        y += 1
        fixed_element = element
        # testing for special case where element starts with '=' which causes error as EXCEL believes it is a formula. 
        if element != "" and element[0] == "=":
            fixed_element = "'" + fixed_element
        ws.cell(row = x, column = y, value = fixed_element)

# Adjusting the column width 
for colname in ['A', 'B', 'C']:
    maxsize = 0
    for cell in ws[colname]:
        maxsize = max(maxsize, len(cell.value))
    ws.column_dimensions[colname].width = maxsize + 5

# Formatting the data in the cells

for cell in ws['B']:
    try: 
        pos = cell.value.rfind('.')         # Checking for extra '.' in the strings.
        if pos>0:
            cell.value = cell.value[:pos-1] # Removing extra '.' in the strings.
        cell.value = int(cell.value)        # Converting STR to INT
        cell.number_format = u'_($* #,##0_)'# Formatting the INT to currency

    except: 
        print()

for cell in ws['C']:
    try: 
        cell.value = int(cell.value)        # Converting the STR to INT
    except: 
        print()
    cell.number_format = u'_(#,##0_)'       # Formatting the INT to numbers

# Inserting the header row
ws.insert_rows(1)
header = ['Country Name', 'External Debt', 'Population']

for y in range(0,len(header)):
    ws.cell(row=1, column=y+1, value = header[y])
    mycell = ws.cell(row=1, column=y+1)
    mycell.font = Font(bold=True)
    mycell.alignment = Alignment(horizontal="center", vertical="center")
    mycell.fill = PatternFill("solid", fgColor="DDDDDD")

# Saving to excel file. 
wb.save('UselessWorldBank.xlsx')

# Clean-up. Closing connections
print("Wrapping up.\n")
conn.close()

# Close final 
with open(logname, 'a') as f:
    f.write(f'{datetime.now().strftime("%y-%m-%d %X")} Ending Daily grind program.\n')

# Calculating elapsed time. Notice the float to 2 decimal place formatting
print(f"Total runtime was: {(time.time()-start_timer):.2f} seconds\n")