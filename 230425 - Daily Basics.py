#############
# Daily Grind

import csv, sqlite3, time
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def english_ord(n):
    """ return string of n followed by the appropriate English Ordinal Suffix"""
    return str(n) + ("th" if ((n%100) in [11,12,23]) else (["th", "st", "nd", "rd"]+["th"]*7)[n%10])

def importdata(dbname, tablename, csvname, logname, label):
    """ Function to import csv file into a DB """    

    write_log(logname, f"Validating if {label} table exists in the database.")
    print(f"Validating if {label} table exists in the database.")

    # Get list of tables from DB, create them if they are not there.
    
    write_log(logname, f"Opening DB: {dbname}")
    conn = sqlite3.Connection(dbname) # open DB Connection
    cursor = conn.cursor()   
  
    write_log(logname, f"Checking if table {tablename} already exists in the DB")
    tables = cursor.execute("SELECT name from sqlite_schema").fetchall()

    if tables == [] or (tablename,) not in tables:      
        # Having confirmed that the table is not there, proceed with importing the csv file
        with open(csvname) as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=',')
            csvDB = tuple(csv_reader) # csv_reader is an iterator. csvDB is not a tuple.
        write_log(logname, f"Reading CSV file: {csvname}")

        fieldnames = csvDB[0]       # Get csv file info: # of fields, header names, etc.
        sqlstring = f"CREATE TABLE {tablename} ("            # Beginning of SQL CREATE command
        for field in fieldnames:
            sqlstring += " '" + field + "' TEXT,"         # Add the various field to create
   
        sqlstring = sqlstring[:-1] + ")"                # Remove last comma and close the bracket
        cursor.execute(sqlstring)
        
        write_log(logname, f"Creating {label} table into the database.")
        print(f"Creating {label} table into the database.")

        ###############################################
        # Populate the database with data from CSV file

        write_log(logname, f"Populating {label} table into the database.\n")
        print(f"Populating {label} table into the database.\n")

        fieldnamestring = "'" + "', '".join(fieldnames) + "'"
        nbfields = len(fieldnames)
        for row in csvDB[1:]:
            sqlstring = f"INSERT INTO {tablename} ({fieldnamestring}) VALUES ({('?,' * nbfields)[:-1]})"
            cursor.execute(sqlstring, row)
    conn.commit()
    conn.close()
    write_log(logname, f"{label} table was created and populated into the database {dbname}.")

def write_log (filename, msg):
    """ This function writes a msg to a log file"""
    with open(filename, 'a') as f:
        f.write(f'{datetime.now().strftime("%y-%m-%d %X")} {msg}\n')


start_time = time.time()

# Generating temporary file name from today's date
basename = "Useless"+date.today().strftime("%y%m%d")
logname = basename + ".log"
dbname = basename + ".db"
xlname = basename + ".xlsx"

write_log(logname, "Begin processing...")

# Print greeting and today's date with correct english ordinal for day and day number.
print(f"Welcome! Today is: {date.today().strftime(f'%A %B ')}"+
                            english_ord(int(date.today().strftime('%d')))+
                            f", the {english_ord(int(date.today().strftime('%j')))}"+
                            f" day of {date.today().strftime('%Y')}"
                            )

importdata(dbname, "products", "Datasets/ShopDB/products.csv", logname, "PRODUCTS")
importdata(dbname, "orders", "Datasets/ShopDB/orders.csv", logname, "ORDERS")
importdata(dbname, "sales", "Datasets/ShopDB/sales.csv", logname, "SALES")
importdata(dbname, "customers", "Datasets/ShopDB/customers.csv", logname,  "CUSTOMERS")

write_log(logname, "Finished importing data...")

#########
# Perform data treatment here...

# open DB Connection
conn = sqlite3.Connection(dbname)
cursor = conn.cursor()

sqlstring = "SELECT * FROM customers"
data = cursor.execute(sqlstring).fetchall()

# getting headerrow information
headerrow = [description[0] for description in cursor.description]
print(headerrow)
write_log(logname, "Processing data...")

#####################################
# Write data back to excel file/sheet
#
write_log(logname, "Saving Data into an Excel workbook.")

print("Saving Data into an Excel workbook.")
wb = Workbook()
ws = wb.active
ws.title = "Shop Dataset"

#
write_log(logname, "Writing Data into an Excel sheet.")
x = 0
for line in data:
    x += 1
    y = 0
    for element in line:
        y += 1
        fixed_element = element
        # testing for special case where element starts with '=' which causes error as EXCEL believes it is a formula. 
        if element != "" and element[0] == "=":
            fixed_element = "'" + fixed_element
        ws.cell(row = x, column = y, value = fixed_element)

# Adjusting the column width 
write_log(logname, "Formatting Excel sheet column width.")

# fix this next line... for column > 7
for colname in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K']:
    maxsize = 0
    for cell in ws[colname]:
        if cell.value !=None:
            maxsize = max(maxsize, len(cell.value))
    ws.column_dimensions[colname].width = maxsize + 5

# Inserting the header row
ws.insert_rows(1)
for y in range(0,len(headerrow)):
    ws.cell(row=1, column=y+1, value = headerrow[y])
    mycell = ws.cell(row=1, column=y+1)
    mycell.font = Font(bold=True)
    mycell.alignment = Alignment(horizontal="center", vertical="center")
    mycell.fill = PatternFill("solid", fgColor="DDDDDD")

write_log(logname, "Inserting and formatting header row.")

# Saving to excel file. 
wb.save(xlname)

write_log(logname, "Saving Excel file to disk.")

# Clean-up. Closing connections
write_log(logname, "Wrapping up...")

print("Wrapping up.\n")
conn.close()

# Close final 
with open(logname, 'a') as f:
    f.write(f'{datetime.now().strftime("%y-%m-%d %X")} Ending Daily grind program.\n')

# Calculating elapsed time. Notice the float to 2 decimal place formatting
final_msg = f"Total runtime was: {(time.time()-start_time):.2f} seconds\n"
print(final_msg)

write_log(logname, final_msg)
