#############
# Daily Grind
#
# ToDo: Format resulting xlsx file. number format, etc...
# ToDo: Try sums, group by, etc. in the SQL query
# ToDo: Improve the readability of the BIG SQL query

import csv, sqlite3, time
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

def english_ord(n):
    """ return string of n followed by the appropriate English Ordinal Suffix"""
    return str(n) + ("th" if ((n%100) in [11,12,23]) else (["th", "st", "nd", "rd"]+["th"]*7)[n%10])

def write_log (filename, msg):
    """ This function writes a msg to a log file and prints it on screen. """
    dated_msg = f'{datetime.now().strftime("%y-%m-%d %X")} {msg}'
    with open(filename, 'a') as f:
        f.write(dated_msg)
    print(dated_msg)
    
def importdata(dbname, tablename, csvname, logname, label):
    """ Function to import csv file into a DB """    

    write_log(logname, f"Validating if {label} table exists in the database.")

    # Get list of tables from DB, create them if they are not there.
    write_log(logname, f"Opening DB: {dbname}")
    conn = sqlite3.Connection(dbname) # open DB Connection
    cursor = conn.cursor()   
  
    write_log(logname, f"Checking if table {tablename} already exists in the DB")
    tables = cursor.execute("SELECT name from sqlite_schema").fetchall()

    if tables == [] or (tablename,) not in tables:      
        # Having confirmed that the table is not there, proceed with importing the csv file
        with open(csvname) as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=',')
            csvDB = tuple(csv_reader) # csv_reader is an iterator. csvDB is not a tuple.
        write_log(logname, f"Reading CSV file: {csvname}")

        fieldnames = csvDB[0]       # Get csv file info: # of fields, header names, etc.
        sqlstring = f"CREATE TABLE {tablename} ("            # Beginning of SQL CREATE command
        for field in fieldnames:
            sqlstring += " '" + field + "' TEXT,"         # Add the various field to create
   
        sqlstring = sqlstring[:-1] + ")"                # Remove last comma and close the bracket
        cursor.execute(sqlstring)
        
        write_log(logname, f"Creating {label} table into the database.")
        print(f"Creating {label} table into the database.")

        ###############################################
        # Populate the database with data from CSV file

        write_log(logname, f"Populating {label} table into the database.")
        fieldnamestring = "'" + "', '".join(fieldnames) + "'"
        nbfields = len(fieldnames)
        for row in csvDB[1:]:
            sqlstring = f"INSERT INTO {tablename} ({fieldnamestring}) VALUES ({('?,' * nbfields)[:-1]})"
            cursor.execute(sqlstring, row)

    conn.commit()
    conn.close()
    write_log(logname, f"{label} table was created and populated into the database {dbname}.")

start_time = time.time()

# Generating temporary file name from today's date
basename = "Useless"+date.today().strftime("%y%m%d")
logname = basename + ".log"
dbname = basename + ".db"
xlname = basename + ".xlsx"

write_log(logname, "Begin processing...")

# Print greeting and today's date with correct english ordinal for day and day number.
print(f"Welcome! Today is: {date.today().strftime(f'%A %B ')}"+
                            english_ord(int(date.today().strftime('%d')))+
                            f", the {english_ord(int(date.today().strftime('%j')))}"+
                            f" day of {date.today().strftime('%Y')}"
                            )

importdata(dbname, "products", "Datasets/ShopDB/products.csv", logname, "PRODUCTS")
importdata(dbname, "orders", "Datasets/ShopDB/orders.csv", logname, "ORDERS")
importdata(dbname, "sales", "Datasets/ShopDB/sales.csv", logname, "SALES")
importdata(dbname, "customers", "Datasets/ShopDB/customers.csv", logname,  "CUSTOMERS")

write_log(logname, "Finished importing data...")

#########
# Perform data treatment here...

# open DB Connection
conn = sqlite3.Connection(dbname)
cursor = conn.cursor()


# Is there a more elegant/readable way to do this SQL string?
 
sqlstring = """
SELECT t5.customer_name, t5.city, t5.order_id, t5.order_date, t5.delivery_date, t5.price_per_unit, t5.quantity, t5.total_price, t6.product_id, t6.product_type, t6.product_name, t6.description
FROM (SELECT product_id, product_type, product_name, description FROM products) as T6
JOIN (SELECT t4.customer_name, t4.city, t3.order_id, t3.order_date, t3.delivery_date, t3.product_id, t3.price_per_unit, t3.quantity, t3.total_price
			 FROM (SELECT t2.customer_id, t2.order_id, t2.order_date, t2.delivery_date, t1.product_id, t1.price_per_unit, t1.quantity, t1.total_price
							FROM (SELECT order_id, product_id, price_per_unit, quantity, total_price FROM sales) as T1
							JOIN (SELECT customer_id, order_id, order_date, delivery_date from orders) as T2
							ON t1. order_id  = t2.order_id) as T3
			JOIN (SELECT customer_id, customer_name, city FROM customers) as T4
			ON t3.customer_id = t4.customer_id) as T5
ON t5.product_id = t6.product_id 
"""

### This SQL query is an equivalent but not much more readable...
"""
WITH sales_data AS (SELECT order_id, product_id, price_per_unit, quantity, total_price FROM sales) ,
order_data AS (SELECT customer_id, order_id, order_date, delivery_date FROM orders) ,
customer_data AS (SELECT customer_id, customer_name, city FROM customers),
product_data AS (SELECT product_id, product_type, product_name, description FROM products)

SELECT *
FROM (SELECT *
				FROM (SELECT *
								FROM (SELECT * FROM sales_data) as t1
								JOIN (SELECT * FROM order_data) as t2
								ON t1.order_id = t2.order_id) as t3
				JOIN (SELECT * from customer_data) as t4
				ON t3.customer_id = t4.customer_id) as t5
JOIN (SELECT * FROM product_data) as t6
ON t5.product_id = t6.product_id
"""

data = cursor.execute(sqlstring).fetchall()

# getting headerrow information
headerrow = [description[0] for description in cursor.description]
write_log(logname, "Processing data...")

#####################################
# Write data back to excel file/sheet
#
write_log(logname, "Saving Data into an Excel workbook.")

wb = Workbook()
ws = wb.active
ws.title = "Shop Dataset"

#
write_log(logname, "Writing Data into an Excel sheet.")
x = 0
for line in data:
    x += 1
    y = 0
    for element in line:
        y += 1
        fixed_element = element
        # testing for special case where element starts with '=' which causes error as EXCEL believes it is a formula. 
        if element != "" and element[0] == "=":
            fixed_element = "'" + fixed_element
        ws.cell(row = x, column = y, value = fixed_element)

# Adjusting the column width 
write_log(logname, "Formatting Excel sheet column width.")

# Inserting the header row
ws.insert_rows(1)
for y in range(0,len(headerrow)):
    ws.cell(row=1, column=y+1, value = headerrow[y])
    mycell = ws.cell(row=1, column=y+1)
    mycell.font = Font(bold=True)
    mycell.alignment = Alignment(horizontal="center", vertical="center")
    mycell.fill = PatternFill("solid", fgColor="DDDDDD")

# Resize columns ['A', 'B', ... ] for the length of headerrow. 
for colname in [chr(i) for i in range(65,91)][:len(headerrow)]:
    maxsize = 0
    for cell in ws[colname]:
        maxsize = max(maxsize, len(cell.value))
    ws.column_dimensions[colname].width = maxsize + 5

write_log(logname, "Inserting and formatting header row.")

# Saving to excel file. 
wb.save(xlname)

write_log(logname, "Saving Excel file to disk.")

# Clean-up. Closing connections
write_log(logname, "Wrapping up...")
conn.close()

# Calculating elapsed time. Notice the float to 2 decimal place formatting
final_msg = f"Ending Daily grind program. Total runtime was: {(time.time()-start_time):.2f} seconds\n"
write_log(logname, final_msg)
